#!/usr/bin/env python3
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import time
from keyboard import press
import openpyxl
from openpyxl import load_workbook
import string
import tkinter as tk
from tkinter import *
import sys

class MyWindow:
    def __init__(self, win):
        #Labels
        self.excelLabel=Label(win, text='Enter Excel Sheet Path')
        self.userLabel=Label(win, text='Username')
        self.passLabel=Label(win, text='Password')
        self.depLabel=Label(win, text='Enter deployment number')
        self.firstLabel=Label(win, text='Enter first Cell')
        self.lastLabel=Label(win, text='Enter Last Cell')
        
        #Entrys
        self.excelEntry=Entry()
        self.userEntry=Entry()
        self.passEntry=Entry(show="*")
        self.depEntry=Entry()
        self.firstEntry=Entry()
        self.lastEntry=Entry()

        #Placement
        self.excelLabel.place(x=100, y=50)
        self.excelEntry.place(x=100, y=70)
        
        self.userLabel.place(x=100, y=100)
        self.userEntry.place(x=100, y=120)
        
        self.passLabel.place(x=100, y=150)
        self.passEntry.place(x=100, y=170)

        self.depLabel.place(x=100, y=200)
        self.depEntry.place(x=100, y=220)

        self.firstLabel.place(x=100, y=250)
        self.firstEntry.place(x=100, y=270)

        self.lastLabel.place(x=100, y=300)
        self.lastEntry.place(x=100, y=320)


        #Buttons
        self.btn1 = Button(win, text='Submit')
        self.b1=Button(win, text='Submit',command=self.change)
        self.b1.place(x=100, y=400)

    def change(self):                       
        username=self.userEntry.get()
        password=self.passEntry.get()
        excelPath=self.excelEntry.get()
        depNum=self.depEntry.get()
        firstCellStr=self.firstEntry.get()
        lastCellStr=self.lastEntry.get()
        firstCell=int(firstCellStr)
        lastCell=int(lastCellStr)+1

        path = excelPath
        driver = webdriver.Chrome(ChromeDriverManager().install())
        driver.get('https://selfservice.denaliai.com/depot')

        serialNumber = '//*[@id="serialbarcode"]'
        binNumber = '//*[@id="binbarcode"]'

        userForm = '//*[@id="username"]'
        passForm = '//*[@id="password"]'
        loginButton = '//*[@id="doLogin"]'
        depotDropdown = '//*[@id="depot"]'
        planoSelect = '//*[@id="depot"]/option[3]'
        placeBin = '//*[@id="linksDiv"]/ul/li[2]/a'

        userName = username
        passWord = password
        depNumber =depNum

        time.sleep(3)
        driver.find_element_by_xpath(userForm).click()
        driver.find_element_by_xpath(userForm).send_keys(userName)

        driver.find_element_by_xpath(passForm).click()
        driver.find_element_by_xpath(passForm).send_keys(passWord)
        driver.find_element_by_xpath(loginButton).click()

        time.sleep(12)
        driver.find_element_by_xpath(depotDropdown).click()

        time.sleep(2)
        driver.find_element_by_xpath(planoSelect).click()


        time.sleep(2)
        driver.find_element_by_xpath(placeBin).click()

        time.sleep(3)

        #loop through serial numbers in excel
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        
        for row in range(firstCell, lastCell):
            cell = sheet_obj.cell(row,column = 2).value
            print(row, " " ,  cell)
            driver.find_element_by_xpath(serialNumber).send_keys(cell)
            time.sleep(1)
            driver.find_element_by_xpath(binNumber).send_keys(depNumber)
            time.sleep(1)
            press('enter')
            time.sleep(2)


window=Tk()
mywin=MyWindow(window)
window.title('Serial To Bin')
window.geometry("500x500+10+10")
window.mainloop()

print('done')



input('Press ENTER to exit')


